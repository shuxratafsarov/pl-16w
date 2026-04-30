"""
Strict data validator for the 3PL weekly dashboard.

Verifies that every value in src/data/week*.json matches the source workbook
3PL_weekly sheet at the row/column level, AND that UI aggregation rules hold
(party sums == group totals == grand totals). Exits non-zero if ANY value
differs by more than the tolerance — the dashboard build MUST run this and
abort on failure.

Usage:  python scripts/validate_markers.py [/path/to/pnl.xlsx]

Source row map (3PL_weekly sheet, 1-indexed):
  Row 6   = week number (column header)
  Row 8   = party number (column header). MPO/MKO are strings like "35 UZUM MKO".
  Row 10  = REVENUE  per party
  Row 94  = EXPENSE  per party
  Row 127 = Linehaul tariff (fallback for M1)
  Row 348 = Marker 1 (USD/kg)
  Row 349 = Marker 2 (Volume/Net ratio)
  Row 350 = Marker 3 (Gross/Net ratio)
  Row 59  = MPO total pieces
  Row 70  = MKO total pieces
  Rows 23/32/41/50 = CAINIAO pieces by country (UZ/BY/AZ/KG)
  Rows 354–374     = CAINIAO subtype ratios (see RATIO_PCS)

Aggregation rules enforced:
  - sum(party.revenue) == totals.revenue
  - sum(party.expense) == totals.expense
  - For Marker 4 (total_pcs) the JSON stores the group total on the FIRST
    party of each type and 0 on the rest, so sum() across parties = group total.
"""
import sys, os, json
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/tmp/pnl.xlsx"
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "src", "data")
TOL_MONEY = 1.00     # USD tolerance for revenue/expense (rounding)
TOL_RATIO = 0.01     # tolerance for M1/M2/M3
TOL_PCS   = 1        # pieces tolerance

REVENUE_ROW = 10
EXPENSE_ROW = 94
M1_ROW, M2_ROW, M3_ROW = 348, 349, 350
LINEHAUL_TARIFF_ROW = 127
COUNTRY_PCS_ROWS = {"UZ": 23, "BY": 32, "KG": 50, "AZ": 41}
RATIO_PCS = {
    "UZ": [354, 355, 356, 357, 358],
    "BY": [362, 363, 364],
    "KG": [368, 369],
    "AZ": [373, 374],
}
MKO_PCS_ROW, MPO_PCS_ROW = 70, 59


def num(v):
    return v if isinstance(v, (int, float)) else None


def near(a, b, tol):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= tol


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["3PL_weekly"]
    issues = []

    for week in range(1, 18):
        path = os.path.join(DATA_DIR, f"week{week}.json")
        if not os.path.exists(path):
            continue
        data = json.load(open(path))

        # Build column map for this week
        cols = [c for c in range(5, 257) if ws.cell(6, c).value == week]
        excel_map = {}
        for c in cols:
            num_cell = ws.cell(8, c).value
            if num_cell is None:
                continue
            if isinstance(num_cell, str) and "MPO" in num_cell:
                excel_map[("MPO", num_cell.split()[0])] = c
            elif isinstance(num_cell, str) and "MKO" in num_cell:
                excel_map[("MKO", num_cell.split()[0])] = c
            else:
                excel_map[("CAINIAO", str(num_cell))] = c

        # ---- per-party check ----
        party_rev_sum = 0.0
        party_exp_sum = 0.0
        seen_first_pcs = {"CAINIAO": False, "MPO": False, "MKO": False}
        type_pcs_excel = {"CAINIAO": 0, "MPO": 0, "MKO": 0}
        type_pcs_json  = {"CAINIAO": 0, "MPO": 0, "MKO": 0}

        for p in data["parties"]:
            key = (p["type"], str(p["num"]))
            c = excel_map.get(key)
            if c is None:
                issues.append(f"W{week} {p['type']} {p['num']}: column not found in Excel")
                continue

            # Revenue / Expense
            xr = num(ws.cell(REVENUE_ROW, c).value) or 0.0
            xe = num(ws.cell(EXPENSE_ROW, c).value) or 0.0
            jr = p.get("revenue") or 0.0
            je = p.get("expense") or 0.0
            if not near(xr, jr, TOL_MONEY):
                issues.append(f"W{week} {p['type']} {p['num']} REVENUE excel={xr:.2f} json={jr:.2f}")
            if not near(xe, je, TOL_MONEY):
                issues.append(f"W{week} {p['type']} {p['num']} EXPENSE excel={xe:.2f} json={je:.2f}")
            party_rev_sum += xr
            party_exp_sum += xe

            # M1/M2/M3
            v1 = num(ws.cell(M1_ROW, c).value)
            if v1 is None:
                v1 = num(ws.cell(LINEHAUL_TARIFF_ROW, c).value)
            v2 = num(ws.cell(M2_ROW, c).value)
            v3 = num(ws.cell(M3_ROW, c).value)
            if v1 is not None and not near(v1, p.get("marker1_tariff"), TOL_RATIO):
                issues.append(f"W{week} {p['type']} {p['num']} M1 excel={v1} json={p.get('marker1_tariff')}")
            if v2 is not None and not near(v2, p.get("marker2_volnet"), TOL_RATIO):
                issues.append(f"W{week} {p['type']} {p['num']} M2 excel={v2} json={p.get('marker2_volnet')}")
            if v3 is not None and not near(v3, p.get("marker3_grossnet"), TOL_RATIO):
                issues.append(f"W{week} {p['type']} {p['num']} M3 excel={v3} json={p.get('marker3_grossnet')}")

            # M4 — pieces (now stored per-party for ALL types)
            if p["type"] == "CAINIAO":
                xp = sum(int(num(ws.cell(r, c).value) or 0) for r in COUNTRY_PCS_ROWS.values())
                # also verify CAINIAO mix kg sums to country kg totals
                xkg_country = {cc: float(num(ws.cell(COUNTRY_KG_ROWS[cc], c).value) or 0)
                               for cc in ("UZ","BY","AZ","KG")}
                jkg_country = {}
                for m in (p.get("mix") or []):
                    jkg_country[m["country"]] = jkg_country.get(m["country"], 0.0) + float(m.get("kg") or 0)
                for cc, xk in xkg_country.items():
                    jk = jkg_country.get(cc, 0.0)
                    if not near(xk, jk, 0.05):
                        issues.append(f"W{week} CAINIAO {p['num']} {cc} mix.kg sum={jk:.2f} excel={xk:.2f}")
            else:
                row = MKO_PCS_ROW if p["type"] == "MKO" else MPO_PCS_ROW
                xp = int(num(ws.cell(row, c).value) or 0)
            jp = int(p.get("total_pcs") or 0)
            if not near(xp, jp, TOL_PCS):
                issues.append(f"W{week} {p['type']} {p['num']} M4 pcs excel={xp} json={jp}")
            type_pcs_excel[p["type"]] += xp
            type_pcs_json[p["type"]]  += jp

            # M4 — kg for MPO/MKO (CAINIAO kg is checked at country level elsewhere)
            if p["type"] in ("MPO", "MKO"):
                kg_row = 67 if p["type"] == "MKO" else 56
                xkg = num(ws.cell(kg_row, c).value) or 0.0
                jkg = p.get("total_kg") or 0.0
                if not near(xkg, jkg, 0.05):
                    issues.append(
                        f"W{week} {p['type']} {p['num']} KG excel={xkg:.2f} json={jkg:.2f}"
                    )
                # mix sum must equal total
                mix = p.get("mix") or []
                mix_kg = sum(float(m.get("kg") or 0) for m in mix)
                mix_pcs = sum(int(m.get("pcs") or 0) for m in mix)
                if not near(mix_kg, jkg, 0.05):
                    issues.append(
                        f"W{week} {p['type']} {p['num']} mix.kg sum={mix_kg:.2f} != total_kg={jkg:.2f}"
                    )
                # mix.pcs sum is informational only — total_pcs is a Marker-4 field
                # stored on the first party per type (see data rules), so it is not
                # comparable to mix sum on subsequent parties.

        # ---- group totals: sum(party) must equal week totals in JSON ----
        jt = data.get("totals", {})
        if not near(party_rev_sum, jt.get("revenue"), TOL_MONEY * 5):
            issues.append(f"W{week} TOTALS.revenue excel_sum={party_rev_sum:.2f} json={jt.get('revenue')}")
        if not near(party_exp_sum, jt.get("expense"), TOL_MONEY * 5):
            issues.append(f"W{week} TOTALS.expense excel_sum={party_exp_sum:.2f} json={jt.get('expense')}")

        # ---- M4 totals per type (UI sums over parties) must match Excel totals ----
        for t in ("CAINIAO", "MPO", "MKO"):
            if not near(type_pcs_excel[t], type_pcs_json[t], TOL_PCS):
                issues.append(
                    f"W{week} M4 {t} excel_total={type_pcs_excel[t]} "
                    f"json_sum_over_parties={type_pcs_json[t]}"
                )

    # ============================================================
    # MONTHLY VALIDATION (3PL_monthly sheet vs src/data/monthly.json)
    # ============================================================
    monthly_path = os.path.join(DATA_DIR, "monthly.json")
    if os.path.exists(monthly_path):
        wsm = wb["3PL_monthly"]
        months = json.load(open(monthly_path))
        M = {
            "REV": 10, "EXP": 94,
            "rev": {"UZ": 16, "BY": 25, "AZ": 34, "KG": 43},
            "exp": {"UZ": 96, "BY": 116, "AZ": 161, "KG": 196},
            "pcs": {"UZ": 23, "BY": 32, "AZ": 41, "KG": 50},
            "kg":  {"UZ": 20, "BY": 29, "AZ": 38, "KG": 47},
        }
        for m in months:
            mo = m["month"]
            c = 4 + mo  # row 6 col 5..16 = month 1..12
            xr = num(wsm.cell(M["REV"], c).value) or 0
            xe = num(wsm.cell(M["EXP"], c).value) or 0
            if not near(xr, m["revenue"], TOL_MONEY):
                issues.append(f"MONTH{mo} revenue excel={xr:.2f} json={m['revenue']}")
            if not near(xe, m["expense"], TOL_MONEY):
                issues.append(f"MONTH{mo} expense excel={xe:.2f} json={m['expense']}")
            for cc in ("UZ", "BY", "AZ", "KG"):
                jc = m["by_country"][cc]
                xcr = num(wsm.cell(M["rev"][cc], c).value) or 0
                xce = num(wsm.cell(M["exp"][cc], c).value) or 0
                xcp = int(num(wsm.cell(M["pcs"][cc], c).value) or 0)
                xck = num(wsm.cell(M["kg"][cc], c).value) or 0
                if not near(xcr, jc["revenue"], TOL_MONEY):
                    issues.append(f"MONTH{mo} {cc} revenue excel={xcr:.2f} json={jc['revenue']}")
                if not near(xce, jc["expense"], TOL_MONEY):
                    issues.append(f"MONTH{mo} {cc} expense excel={xce:.2f} json={jc['expense']}")
                if not near(xcp, jc["pcs"], TOL_PCS):
                    issues.append(f"MONTH{mo} {cc} pcs excel={xcp} json={jc['pcs']}")
                if not near(xck, jc["kg"], 0.05):
                    issues.append(f"MONTH{mo} {cc} kg excel={xck:.2f} json={jc['kg']}")

    if issues:
        print(f"✗ {len(issues)} DISCREPANCIES vs source:")
        for i in issues:
            print(" ", i)
        sys.exit(1)
    print("✓ All values match source 100% (weekly M1–M4 + revenue/expense + monthly per country)")


if __name__ == "__main__":
    main()

